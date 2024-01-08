from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.core.window import Window
from kivymd.uix.dialog import MDDialog
from kivy.app import App
import openpyxl
from openpyxl import Workbook
import pathlib
from kivymd.uix.datatables import MDDataTable
from kivy.metrics import dp
from kivymd.uix.button import MDRaisedButton

KV='''
Screen:
    MDBoxLayout:
        orientation: 'vertical'
        padding: 16
        spacing:5

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3

            MDTextField:
                hint_text: "Name"
                mode: "rectangle"
                id: name
                icon_right: "account-outline"
                width: 200
                font_size: 48
                pos_hint: {"center_x": 0.5}
                size_hint_y: 1.2
                size_hint_x: 2

            MDTextField:
                hint_text: "Standard"
                mode: "rectangle"
                id: standard
                icon_right: "list-box-outline"
                width: 200
                font_size: 48
                size_hint_y: 1.2
                size_hint_x: 2

            MDRaisedButton:
                text: "Add"
                md_bg_color: "#333333"
                size_hint_y: 1.08
                on_press: app.add()

        MDCard:
            size_hint: None, None
            size: 1005, 1700
            pos_hint: {'center_x': 0.5, 'center_y': 0.5}
            padding: 15
            spacing: 25

            MDScrollView:
                id: card_layout
                do_scroll_x: False

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3
            MDRaisedButton:
                text: "Search"
                icon_right: "database-search-outline"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.search()
            MDRaisedButton:
                id: reset_button
                text: "Reset Table"
                icon_right: "refresh"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.reset_table()
'''


class dataEntry(MDApp):
    def build(self):
        # Your build method code here
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "DeepPurple"
        return Builder.load_string(KV)

    def add(self, *args):
        # Your add method code here
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' or standard == '':
            print("Fill all fields")
        else:
            print(name)
            print(standard)

            file = openpyxl.load_workbook('Backend_data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=standard)

            file.save(r'Backend_data.xlsx')
            self.reset_table()

            self.root.ids.name.text = ''
            self.root.ids.standard.text = ''

    def reset_search_buttons(self, search_button_visible=True):
        # Your reset_search_buttons method code here
        self.root.ids.search_button.opacity = 1 if search_button_visible else 0
        self.root.ids.reset_button.opacity = 1 if not search_button_visible else 0

    def search(self):
        # Your search method code here
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' and standard == '':
            print("Enter at least one value for search")
        else:
            data = self.read_excel('Backend_data.xlsx')

            if name:
                data = [row for row in data if name.lower() in row[0].lower()]
            if standard:
                data = [row for row in data if standard.lower() in str(row[1]).lower()]

            cols = ['Name', 'Standard']
            values = data

            table = MDDataTable(
                pos_hint={'center_x': 0.2, 'center_y': 0.2},
                column_data=[(col, dp(35)) for col in cols],
                row_data=values,
                use_pagination=True,
                padding=10,
                pagination_menu_pos='auto',
                rows_num=10,
            )

            self.root.ids.card_layout.clear_widgets()
            self.root.ids.card_layout.add_widget(table)

    def on_start(self):
        # Your on_start method code here
        file = pathlib.Path('Backend_data.xlsx')
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Name"
            sheet['B1'] = "Standard"
            file.save('Backend_data.xlsx')

        self.reset_table()

    def reset_table(self):
        # Your reset_table method code here
        data = self.read_excel('Backend_data.xlsx')
        cols = ['Name', 'Standard']
        values = data

        table = MDDataTable(
            pos_hint={'center_x': 0.2, 'center_y': 0.2},
            column_data=[(col, dp(35)) for col in cols],
            row_data=values,
            use_pagination=True,
            padding=10,
            pagination_menu_pos='auto',
            rows_num=10,
        )
        self.root.ids.card_layout.clear_widgets()
        self.root.ids.card_layout.add_widget(table)

    def read_excel(self, filename):
        # Read data from Excel file without Pandas
        file = openpyxl.load_workbook(filename)
        sheet = file.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data

if __name__ == "__main__":
    data_app = dataEntry()
    data_app.run()
